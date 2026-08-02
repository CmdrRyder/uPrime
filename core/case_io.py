"""
core/case_io.py
---------------
Reader and data model for uPrime's OWN 1D exports, used by the standalone
"Compare Cases" viewer (Stage A, 1D only).

This module never reads raw u/v datasets. It reads the files uPrime itself
writes:

  * 1D line profiles  -> CSV via core.export.export_line_csv
        columns: dist_mm, x_mm, y_mm, mean_<name>, [std_<name>], ...
  * 1D spectra        -> CSV via core.export.export_spectra_csv
        columns: frequency_Hz, PSD_<comp>_m2s2_per_Hz, ...
  * 2D fields         -> Tecplot ASCII .dat via core.export.export_2d_tecplot
        (VARIABLES = ... / ZONE ... F=POINT).  Detected but NOT loaded in
        Stage A.

Every 1D export starts with the shared "#"-comment settings header emitted by
core.export._settings_header:

    # uPrime Export
    # Generated : YYYY-mm-dd HH:MM:SS
    # ==================================================
    # Analysis                 : <module>
    # <key>                    : <value>
    # ==================================================
    <col1>,<col2>,...
    <data rows>

The reader is version-safe (plain numpy) and NaN-safe (blank / non-numeric
cells become NaN and pass through untouched).
"""

import os
import re
from dataclasses import dataclass, field
from typing import Optional

import numpy as np


# Point-coordinate columns that accompany a line profile but are not
# themselves plottable quantities (they describe the sampling path).
COORD_COLS = {"x_mm", "y_mm"}

# x-type -> (default axis symbol, default units)
X_TYPE_META = {
    "arc_length": ("s", "mm"),
    "frequency":  ("f", "Hz"),
    "wavenumber": ("k", "rad/m"),
}

# Canonical quantities and the fixed axis metadata they carry.  Anything not
# listed here falls back to its raw column label as the key, taking its x-axis
# from the file (so unknown modules still plot).
#   x: (x_type, x_label, x_units)
_ARC = ("arc_length", "s", "mm")
QUANTITY_INFO = {
    "U":    {"x": _ARC, "y_label": "U",       "y_units": "m/s"},
    "V":    {"x": _ARC, "y_label": "V",       "y_units": "m/s"},
    "W":    {"x": _ARC, "y_label": "W",       "y_units": "m/s"},
    # ASCII angle brackets to match reynolds_window's COMP_LABELS and stay
    # renderable in the app's serif font (Times New Roman lacks U+27E8/9).
    "R_uu": {"x": _ARC, "y_label": "<u'u'>", "y_units": "m²/s²"},
    "R_vv": {"x": _ARC, "y_label": "<v'v'>", "y_units": "m²/s²"},
    "R_ww": {"x": _ARC, "y_label": "<w'w'>", "y_units": "m²/s²"},
    "R_uv": {"x": _ARC, "y_label": "<u'v'>", "y_units": "m²/s²"},
    "R_uw": {"x": _ARC, "y_label": "<u'w'>", "y_units": "m²/s²"},
    "R_vw": {"x": _ARC, "y_label": "<v'w'>", "y_units": "m²/s²"},
    "TKE":  {"x": _ARC, "y_label": "k",       "y_units": "m²/s²"},
}


# Components whose values are signed (use a diverging colormap when displayed).
SIGNED_COMPONENTS = {"V", "W", "R_uv", "R_uw", "R_vw"}


def is_signed_component(key):
    """True if a component is known to be signed (RdBu_r); False if positive."""
    return key in SIGNED_COMPONENTS


def _norm_stress(base):
    """Reduce a stress-like token (uu, <u'u'>, vu, ...) to canonical uu/uv/... ."""
    comp = re.sub(r"[^uvw]", "", base)
    if len(comp) == 2:
        canon = "".join(sorted(comp))          # vu -> uv, uu -> uu
        if canon in ("uu", "vv", "ww", "uv", "uw", "vw"):
            return canon
    return None


def component_key_from_varname(name):
    """Map a Tecplot VARIABLES field name to a canonical component key.

    Handles the unit suffix ("U [m/s]" -> U), Reynolds-stress angle-bracket
    tokens ("<u'u'> [m2/s2]" -> R_uu, even when scaled by Um), and the common
    magnitude / TKE labels; falls back to the raw (unit-stripped) name.
    """
    base = re.sub(r"\s*\[.*?\]\s*$", "", name).strip()   # drop trailing [units]
    m = re.search(r"<([^>]*)>", base)                    # <u'u'> stress token
    if m:
        st = _norm_stress(m.group(1))
        if st:
            return "R_" + st
    key = detect_quantity_key(base)
    if key == base:                                      # nothing matched
        low = base.lower()
        if low in ("mag", "|v|", "speed", "vmag"):
            return "MAG"
        if "tke" in low or low == "k":
            return "TKE"
    return key


def detect_quantity_key(col, meta=None):
    """Map a uPrime export column name to a canonical quantity_key.

    Explicit mappings for the common line-profile quantities; everything else
    (budget terms, spectra PSD, std columns, unknown modules) falls back to the
    raw column label so it still plots under its own variable.
    """
    name = col.strip()
    low = name.lower()

    # std_* columns keep their own identity (paired stds, not the mean quantity)
    if low.startswith("std_"):
        return name
    # spectra / unit-laden columns are never Reynolds stresses
    if any(tok in low for tok in ("psd", "per_hz", "_hz")):
        return name

    base = low[5:] if low.startswith("mean_") else low
    stress = _norm_stress(base)
    if stress:
        return "R_" + stress
    if base in ("u", "v", "w"):
        return base.upper()
    if base in ("k", "tke"):
        return "TKE"
    return name


# --------------------------------------------------------------------------- #
# Data model
# --------------------------------------------------------------------------- #

@dataclass
class CaseSeries:
    """One plottable column from a loaded uPrime 1D export."""
    source_file:   str
    source_module: str
    quantity_key:  str            # canonical quantity (U, R_uu, TKE, ...)
    x_type:        str
    x_label:       str
    x_units:       str
    x_data:        np.ndarray
    y_label:       str
    y_units:       str
    y_data:        np.ndarray
    um:            Optional[float]   # None when not applicable
    label:         str            # editable display label
    style:         dict           # {"color", "linestyle", "marker"}
    enabled:       bool = True
    source_kind:   str = "csv"    # "csv" (imported) or "extracted" (2D line)


@dataclass
class FileRecord:
    """
    Result of reading one file.

    For a 2D Tecplot field, ``is_2d`` is True and ``series`` is empty (Stage A
    lists it but does not plot it).  For a 1D tabular export, ``series`` holds
    one CaseSeries per numeric non-x column once the x-axis is known.  If the
    x-axis type could not be detected, ``x_type == "unknown"`` and ``series``
    stays empty until :meth:`set_x_column` is called.
    """
    source_file:   str
    path:          str
    is_2d:         bool = False
    source_module: Optional[str] = None
    x_type:        str = "unknown"
    x_label:       str = ""
    x_units:       str = ""
    x_col:         Optional[str] = None
    columns:       list = field(default_factory=list)
    series:        list = field(default_factory=list)
    twod_quantities: list = field(default_factory=list)  # field names in a 2D .dat
    twod_components: list = field(default_factory=list)   # canonical keys of those
    error:         Optional[str] = None
    # --- 2D full-field data (loaded lazily via load_field(), Stage B) ---
    x:             object = None          # 1D np.ndarray (mm)
    y:             object = None          # 1D np.ndarray (mm)
    components:    dict = field(default_factory=dict)      # key -> 2D float32 array
    value_ranges:  dict = field(default_factory=dict)      # key -> (vmin, vmax)
    field_loaded:  bool = False
    _table:        dict = field(default_factory=dict, repr=False)
    _meta:         dict = field(default_factory=dict, repr=False)
    _um:           Optional[float] = None

    def load_field(self):
        """Lazily load the full 2D field arrays for a Tecplot record.

        Populates ``x`` (1D mm), ``y`` (1D mm), ``components`` (canonical key ->
        2D float32 array) and ``value_ranges`` (key -> (vmin, vmax)). NaN /
        masked regions are preserved as NaN. Idempotent and a no-op for 1D
        records.
        """
        if self.field_loaded or not self.is_2d:
            return
        x1, y1, comps = _load_tecplot_field(self.path)
        self.x, self.y, self.components = x1, y1, comps
        ranges = {}
        for k, v in comps.items():
            if np.isfinite(v).any():
                ranges[k] = (float(np.nanmin(v)), float(np.nanmax(v)))
            else:
                ranges[k] = (float("nan"), float("nan"))
        self.value_ranges = ranges
        self.field_loaded = True

    def set_x_column(self, col):
        """Manually choose the independent (x) column for an unknown file.

        Used by the GUI when auto-detection was ambiguous. Rebuilds the
        series list against the chosen column.
        """
        if col not in self.columns:
            raise ValueError(f"{col!r} is not a column of {self.source_file}")
        self.x_col   = col
        self.x_type  = f"custom:{col}"
        self.x_label = col
        self.x_units = ""
        self.series  = _build_series(self)
        return self.series


# --------------------------------------------------------------------------- #
# Low-level parsing helpers
# --------------------------------------------------------------------------- #

def _to_float_array(values):
    """Convert a list of strings to a float64 array; blanks / non-numeric -> NaN."""
    out = np.empty(len(values), dtype=np.float64)
    for i, v in enumerate(values):
        if v is None:
            out[i] = np.nan
            continue
        s = str(v).strip()
        if s == "":
            out[i] = np.nan
            continue
        try:
            out[i] = float(s)
        except ValueError:
            out[i] = np.nan
    return out


def _split_row(line, delim):
    if delim == ",":
        return [c.strip() for c in line.split(",")]
    return line.split()


def _parse_comment_meta(comment_lines):
    """Parse '# key : value' metadata lines into a dict (banner lines skipped)."""
    meta = {}
    for raw in comment_lines:
        body = raw.lstrip("#").strip()
        if not body:
            continue
        if body == "uPrime Export" or body.startswith("Generated"):
            continue
        if set(body) <= {"="}:            # separator rule "===="
            continue
        if ":" in body:
            key, val = body.split(":", 1)
            meta[key.strip()] = val.strip()
    return meta


def _looks_like_tecplot(lines):
    """True if any line declares a Tecplot ZONE / VARIABLES header."""
    for ln in lines:
        s = ln.strip().upper()
        if s.startswith("ZONE") or s.startswith("VARIABLES"):
            return True
    return False


def _parse_tecplot_variables(lines):
    """Return the field names from a Tecplot VARIABLES line, excluding x/y."""
    for ln in lines:
        if ln.strip().upper().startswith("VARIABLES"):
            names = re.findall(r'"([^"]*)"', ln)
            return names[2:] if len(names) > 2 else names
    return []


def _load_tecplot_field(path):
    """Fully parse a Tecplot POINT-format .dat -> (x_1d, y_1d, components).

    x_1d, y_1d are 1D ascending physical axes (mm, float32); components maps a
    canonical component key to a 2D float32 array of shape (len(y), len(x)).
    Non-numeric / 'nan' cells are preserved as NaN.
    """
    varnames = []
    nx = ny = None
    rows = []
    with open(path, encoding="utf-8-sig", errors="replace") as fh:
        for line in fh:
            s = line.strip()
            if not s or s.startswith("#"):
                continue
            up = s.upper()
            if up.startswith("TITLE"):
                continue
            if up.startswith("VARIABLES"):
                varnames = re.findall(r'"([^"]*)"', s)
                continue
            if up.startswith("ZONE"):
                mi = re.search(r"\bI\s*=\s*(\d+)", s)
                mj = re.search(r"\bJ\s*=\s*(\d+)", s)
                if mi:
                    nx = int(mi.group(1))
                if mj:
                    ny = int(mj.group(1))
                continue
            rows.append(s.split())

    if not rows:
        raise ValueError("Tecplot file contains no data rows")
    arr = np.array([[_to_float(v) for v in row] for row in rows],
                   dtype=np.float64)
    npts = arr.shape[0]
    if nx is None or ny is None or nx * ny != npts:
        raise ValueError(
            f"Tecplot I/J ({nx}x{ny}) do not match {npts} data points")

    field_names = varnames[2:] if len(varnames) > 2 else []
    Xg = arr[:, 0].reshape(ny, nx)
    Yg = arr[:, 1].reshape(ny, nx)
    flip_x = Xg[0, 0] > Xg[0, -1]
    flip_y = Yg[0, 0] > Yg[-1, 0]

    comps = {}
    for idx, name in enumerate(field_names):
        key = component_key_from_varname(name)
        grid = arr[:, 2 + idx].reshape(ny, nx).astype(np.float32)
        if flip_x:
            grid = grid[:, ::-1]
        if flip_y:
            grid = grid[::-1, :]
        comps[key] = grid

    x1 = Xg[0, :].astype(np.float32)
    y1 = Yg[:, 0].astype(np.float32)
    if flip_x:
        x1 = x1[::-1]
    if flip_y:
        y1 = y1[::-1]
    return np.ascontiguousarray(x1), np.ascontiguousarray(y1), comps


def _to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return np.nan


def _read_delimited(path):
    """Read a uPrime CSV/DAT 1D export -> (meta, columns, table{name: float array})."""
    with open(path, newline="", encoding="utf-8-sig") as fh:
        raw_lines = fh.read().splitlines()

    comment_lines = []
    header = None
    data_lines = []
    for ln in raw_lines:
        s = ln.strip()
        if not s:
            continue
        if s.startswith("#"):
            comment_lines.append(s)
            continue
        if header is None:
            header = ln
        else:
            data_lines.append(ln)

    if header is None:
        raise ValueError("No column header row found")

    delim = "," if "," in header else None       # None -> whitespace split
    columns = _split_row(header, delim)

    raw_cols = {c: [] for c in columns}
    for ln in data_lines:
        parts = _split_row(ln, delim)
        for i, c in enumerate(columns):
            raw_cols[c].append(parts[i] if i < len(parts) else "")

    meta = _parse_comment_meta(comment_lines)
    table = {c: _to_float_array(v) for c, v in raw_cols.items()}
    return meta, columns, table


def _read_xlsx(path):
    """Read a uPrime-style .xlsx 1D export -> (meta, columns, table).

    uPrime does not currently write .xlsx, but the reader supports it: leading
    rows whose first cell starts with '#' are treated as the metadata header
    (same convention as the CSV exports), the next row is the column header,
    and the remaining rows are data.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "Reading .xlsx requires the 'openpyxl' package.") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    comment_lines = []
    header = None
    data_rows = []
    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        first = row[0]
        if header is None and isinstance(first, str) and first.lstrip().startswith("#"):
            comment_lines.append(str(first))
            continue
        if all(c is None for c in row):
            continue
        if header is None:
            header = [("" if c is None else str(c)).strip() for c in row]
        else:
            data_rows.append(row)
    wb.close()

    if header is None:
        raise ValueError("No column header row found in xlsx")

    columns = [c for c in header if c != ""]
    raw_cols = {c: [] for c in columns}
    for row in data_rows:
        for i, c in enumerate(columns):
            raw_cols[c].append(row[i] if i < len(row) else None)

    meta = _parse_comment_meta(comment_lines)
    table = {c: _to_float_array(v) for c, v in raw_cols.items()}
    return meta, columns, table


# --------------------------------------------------------------------------- #
# Detection
# --------------------------------------------------------------------------- #

def _detect_x(columns, meta):
    """Return (x_col, x_type, x_label, x_units) from columns + metadata.

    Falls back to (None, "unknown", "", "") when the axis cannot be inferred,
    so the GUI can prompt for a manual x-column override.
    """
    lower = {c.lower(): c for c in columns}
    analysis = (meta.get("Analysis") or "").lower()

    # Line profiles carry an arc-length / distance column.
    if "dist_mm" in lower:
        sym, units = X_TYPE_META["arc_length"]
        return lower["dist_mm"], "arc_length", sym, units

    # Spectra write "frequency_Hz" as the first column for BOTH temporal and
    # spatial exports; disambiguate the semantic axis from the header.
    if "frequency_hz" in lower:
        is_spatial = (
            "spatial" in analysis
            or "wavenumber" in analysis
            or analysis.strip() == "spectral"
        )
        if is_spatial:
            sym, units = X_TYPE_META["wavenumber"]
            return lower["frequency_hz"], "wavenumber", sym, units
        sym, units = X_TYPE_META["frequency"]
        return lower["frequency_hz"], "frequency", sym, units

    # Explicit wavenumber columns.
    for c in columns:
        cl = c.lower()
        if cl in ("k", "wavenumber", "k_rad_m") or cl.startswith("k_rad"):
            sym, units = X_TYPE_META["wavenumber"]
            return c, "wavenumber", sym, units

    return None, "unknown", "", ""


def _extract_um(meta):
    """Recover Um from the metadata header, or None if absent / not applicable."""
    for key in ("U_m", "Um", "Scaled by Um", "Normalized"):
        if key not in meta:
            continue
        val = meta[key]
        if "n/a" in val.lower():
            continue
        m = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", val)
        if m:
            try:
                return float(m.group())
            except ValueError:
                pass
    return None


def _col_units(name):
    """Best-effort y-axis units from a uPrime column name."""
    n = name.lower()
    if "m2s2_per_hz" in n:
        return "m²/s²/Hz"
    return ""


# --------------------------------------------------------------------------- #
# Series construction
# --------------------------------------------------------------------------- #

def _build_series(record):
    """Build one CaseSeries per numeric, non-x, non-coordinate column."""
    series = []
    base = os.path.splitext(record.source_file)[0]
    x_data = record._table[record.x_col]
    module = record.source_module or ""

    for col in record.columns:
        if col == record.x_col or col in COORD_COLS:
            continue
        y_data = record._table[col]
        if not np.isfinite(y_data).any():
            # Fully blank / non-numeric column (e.g. a text "case" column).
            continue

        qkey = detect_quantity_key(col, record._meta)
        info = QUANTITY_INFO.get(qkey)
        if info:
            x_type, x_label, x_units = info["x"]
            y_label, y_units = info["y_label"], info["y_units"]
        else:
            x_type, x_label, x_units = (record.x_type, record.x_label,
                                        record.x_units)
            y_label, y_units = col, _col_units(col)

        series.append(CaseSeries(
            source_file=record.source_file,
            source_module=module,
            quantity_key=qkey,
            x_type=x_type,
            x_label=x_label,
            x_units=x_units,
            x_data=x_data,
            y_label=y_label,
            y_units=y_units,
            y_data=y_data,
            um=record._um,
            label=base,                # default label = filename (stem)
            style={"color": None, "linestyle": "-", "marker": ""},
            enabled=True,
        ))
    return series


# --------------------------------------------------------------------------- #
# Public entry point
# --------------------------------------------------------------------------- #

def read_case_file(path):
    """
    Read one uPrime export file and return a :class:`FileRecord`.

    Type is detected by CONTENT, not extension:
      * a Tecplot ZONE/VARIABLES header  -> is_2d=True (not loaded in Stage A);
      * anything else                    -> 1D tabular export.

    Read/parse failures are captured on ``record.error`` rather than raised, so
    a single bad file does not abort a multi-file load.
    """
    base = os.path.basename(path)
    rec = FileRecord(source_file=base, path=path)
    ext = os.path.splitext(path)[1].lower()

    try:
        if ext == ".xlsx":
            meta, columns, table = _read_xlsx(path)
        else:
            with open(path, encoding="utf-8-sig", errors="replace") as fh:
                head = []
                for _ in range(80):
                    line = fh.readline()
                    if not line:
                        break
                    head.append(line)
            if _looks_like_tecplot(head):
                # Stage A: not rendered, but still parse metadata so the GUI
                # can park it under its source module with its quantity names.
                rec.is_2d = True
                comment_lines = [ln for ln in head if ln.strip().startswith("#")]
                rec._meta = _parse_comment_meta(comment_lines)
                rec.source_module = rec._meta.get("Analysis")
                rec.twod_quantities = _parse_tecplot_variables(head)
                rec.twod_components = [component_key_from_varname(v)
                                       for v in rec.twod_quantities]
                return rec
            meta, columns, table = _read_delimited(path)

        rec._meta = meta
        rec._table = table
        rec.columns = columns
        rec.source_module = meta.get("Analysis")
        rec._um = _extract_um(meta)

        x_col, x_type, x_label, x_units = _detect_x(columns, meta)
        rec.x_col, rec.x_type = x_col, x_type
        rec.x_label, rec.x_units = x_label, x_units
        if x_type != "unknown" and x_col is not None:
            rec.series = _build_series(rec)
    except Exception as exc:                      # noqa: BLE001 - reported to GUI
        rec.error = str(exc)

    return rec
